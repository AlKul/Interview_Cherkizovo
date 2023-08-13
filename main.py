import tkinter as tk
from tkinter import filedialog
import pyodbc
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os
from dotenv import load_dotenv

load_dotenv()

server = os.environ.get("SERVER")
port = os.environ.get("PORT")
database = os.environ.get("DATABASE")
username = os.environ.get("USERNAME")
password = os.environ.get("PASSWORD")


def upload_data():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx"), ("Excel Files", "*.xlsb")])

    if file_path:
        df = pd.read_excel(file_path)
        print(df.shape)
        conn = pyodbc.connect(
            "DRIVER={ODBC Driver 18 for SQL Server};"
            f'SERVER={server};'
            f'PORT={port};'
            f'DATABASE={database};'
            f'UID={username};'
            f'PWD={password};'
            f'Trusted_Connection=yes;'
        )
        cursor = conn.cursor()
        for index, row in df.iterrows():
            cursor.execute("INSERT INTO data (date, articul, sales) VALUES (?,?,?)",
                           row['date'], row['articul'], row['sales'])
            conn.commit()
        cursor.close()
        conn.close()


def download_data():
    conn = pyodbc.connect(f'DRIVER={{ODBC Driver 18 for SQL Server}};'
                          f'SERVER={server};UID={username};PWD={password}')
    cursor = conn.cursor()

    start = start_date.get()
    end = end_date.get()

    cursor.execute("{call GetReport(?, ?)}", start, end)
    data = cursor.fetchall()
    cursor.close()
    conn.close()

    if data:
        book = Workbook()
        sheet = book.active
        headers = ["Year", "Month", "Articul", "Average Sales", "Total Share"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet.column_dimensions[col_letter].width = max(len(header), 12)
            header_cell = sheet.cell(row=1, column=col_num)
            header_cell.value = header
            header_cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"

        for row_num, row_data in enumerate(data, 2):
            for col_num, cell_data in enumerate(row_data, 1):
                col_letter = get_column_letter(col_num)
                sheet.column_dimensions[col_letter].width = max(len(str(cell_data)), 12)
                sheet.cell(row=row_num, column=col_num, value=cell_data)

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if file_path:
            book.save(file_path)


root = tk.Tk()
root.geometry("400x300")
root.title("Sales Data App")

upload_button = tk.Button(root, text="Upload Data", command=upload_data)
upload_button.pack(pady=20)

date_label = tk.Label(root, text="Enter date range:")
date_label.pack()

start_label = tk.Label(root, text="Start date (YYYY-MM-DD):")
start_label.pack()

start_date = tk.Entry(root)
start_date.pack()

end_label = tk.Label(root, text="End date (YYYY-MM-DD):")
end_label.pack()

end_date = tk.Entry(root)
end_date.pack()

download_button = tk.Button(root, text="Download Report", command=download_data)
download_button.pack(pady=20)

root.mainloop()
